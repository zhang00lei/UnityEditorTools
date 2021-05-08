# -*- coding:utf-8 -*-

import sys, os, re
import openpyxl

curpath = os.path.dirname(os.path.abspath(sys.argv[0]))

# 将数据导出到tgt_lua_path
def toInt(val):
    if isinstance(val, str):
        v = int(val)
    else:
        v = int(val)
    return v;

# 将数据导出到tgt_lua_path
def excel2lua(src_excel_path, tgt_lua_path):
    excel_data_src = openpyxl.load_workbook(src_excel_path, data_only=True)
    excel_sheet = excel_data_src.worksheets[0]
    if src_excel_path.endswith('CommonConfig.xlsx'):
        lua_export_file = open(tgt_lua_path, 'w',encoding='utf-8')
        lua_export_file.write('---this file is generate by tools,do not modify it.\n')
        lua_export_file.write('---@class CommonConfig\n')
        lua_export_file.write('local CommonConfig = {}\n')
        for row in range(1, excel_sheet.max_row):
            field_name = excel_sheet.cell(row+1,1).value
            field_note = excel_sheet.cell(row+1,2).value
            field_type = excel_sheet.cell(row+1,3).value   
            field_value = excel_sheet.cell(row+1,4).value
            if field_type == 'float' or field_type == 'int' or field_type == 'number':
                field_type = 'number'
            elif field_type == 'boolean':
                if not field_value:
                    field_value = 'false'
                else:
                    field_value = 'true'

            lua_export_file.write('---@field {0} {1} @{2}\n'.format(field_name, field_type,field_note))
            if field_type == 'string':
                lua_export_file.write("CommonConfig.{0} = \"{1}\"\n".format(field_name,field_value))
            else:
                lua_export_file.write("CommonConfig.{0} = {1}\n".format(field_name,field_value))
        lua_export_file.write('return CommonConfig')
        lua_export_file.close()
        print('exported')
        return

    # excel data dict
    excel_data_dict = {}

    # col desc
    col_desc_list = []

    # col name list
    col_name_list = []

    # col val type list
    col_val_type_list = []

    # 第一行是表名
    # 第二行是所有列的描述
    for col in range(0, excel_sheet.max_column):
        cell = excel_sheet.cell(2, col + 1)
        col_desc_list.append(str(cell.value))

    # 遍历第三行的所有列 保存字段名
    for col in range(0, excel_sheet.max_column):
        cell = excel_sheet.cell(3, col + 1)
        if cell.value:
            col_name_list.append(str(cell.value))
        elif cell.data_type != "s":
            print("found a invalid col name in col [%d] !~" % (col + 1))

    # 遍历第四行的所有列 保存数据类型
    for col in range(0, excel_sheet.max_column):
        cell = excel_sheet.cell(4, col + 1)
        if cell.value:
            col_val_type_list.append(str(cell.value))
        elif cell.data_type != "s":
            print("found a invalid col val type in col [%d] !~" % (col + 1))

    # 剔除表头、字段名和字段类型所在行
    # 从第五行开始遍历 构造行数据
    for row in range(4, excel_sheet.max_row):
        # 保存数据索引 默认第一列为id
        cell_id = excel_sheet.cell(row + 1, 1)

        # assert cell_id.data_type == 2, "found a invalid id in row [%d] !~" % (row)
        # 检查id的唯一性
        if not cell_id.value:
            continue;

        if cell_id.value in excel_data_dict:
            print('[warning] duplicated data id: "%d", all previous value will be ignored!~' % (cell_id.value))

        # row data list
        row_data_list = []

        # 保存每一行的所有数据
        for col in range(0, len(col_name_list)):
            cell = excel_sheet.cell(row + 1, col + 1)
            k = col_name_list[col]
            cell_val_type = col_val_type_list[col]

            # print("row", row, "col", col)
            # ignored the string that start with '_'
            if str(k).startswith('#'):
                continue

            # 根据字段类型去调整数值 如果为空值 依据字段类型 填上默认值
            if cell_val_type == 'string':
                if not cell.value:
                    v = '\"\"'
                else:
                    v = '\"%s\"' % (str(cell.value))
            elif cell_val_type == 'int':
                if not cell.value:
                    v = 0
                else:
                    v = toInt(cell.value)
            elif cell_val_type == 'float':
                if not cell.value:
                    v = 0
                else:
                    v = float(cell.value)
            elif cell_val_type == 'bool':
                if not cell.value:
                    v = 'false'
                else:
                    v = 'true'
            elif cell_val_type == 'table':
                if not cell.value:
                    v = '{}'
                else:
                    v = cell.value
            else:
                v = cell.value

            # 加入列表
            row_data_list.append([k, v])

        # 保存id 和 row data
        excel_data_dict[cell_id.value] = row_data_list

    searchObj = re.search(r'([^\\/:*?"<>|\r\n]+)\.\w+$', tgt_lua_path, re.M | re.I)
    lua_table_name = searchObj.group(1)

    # 这个就直接获取文件名了
    src_excel_file_name = os.path.basename(src_excel_path)
    tgt_lua_file_name = os.path.basename(tgt_lua_path)

    # export to lua file
    lua_export_file = open(tgt_lua_path, 'w',encoding='utf-8')

    lua_export_file.write('---@class {0}\n'.format(lua_table_name))

    for i in range(len(col_name_list)):
        type_name = col_val_type_list[i]
        if col_name_list[i].startswith("#"):
            continue
        if type_name == 'float' or type_name == 'int' or type_name == 'number':
            type_name = 'number'
        type_des = col_desc_list[i].replace('\n', ' ')
        lua_export_file.write('---@field {0} {1} @{2}\n'.format(col_name_list[i], type_name, type_des))

    lua_export_file.write('local %s = {\n' % lua_table_name)

    # 遍历excel数据字典 按格式写入
    for k, v in excel_data_dict.items():
        lua_export_file.write('    [%d] = {\n' % toInt(k))
        for row_data in v:
            lua_export_file.write('        {0} = {1},\n'.format(row_data[0], row_data[1]))
        lua_export_file.write('    },\n')

    lua_export_file.write('}\n')
    lua_export_file.write('return %s' % lua_table_name)

    lua_export_file.close()

    print('[excel] %d row data exported !~ %s' % (excel_sheet.max_row, os.path.basename(tgt_lua_path)))


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('python excel2lua.py <excel_input_path> <lua_output_path>')
        exit(1)

    excel2lua(os.path.join(curpath, sys.argv[1]), os.path.join(curpath, sys.argv[2]))

    # test hero.xlsx
    # curpath = "E:/craftclient/ConfigData/Excel2Lua"
    # excelPath = curpath + "/excel/CommonConfig.xlsx";
    # luaPath = curpath + "/lua/CommonConfig.lua";
    # excel2lua(excelPath, luaPath)

    exit(0)
